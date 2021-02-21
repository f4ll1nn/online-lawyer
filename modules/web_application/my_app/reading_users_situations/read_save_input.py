""" Realization usersituation upworking """
import os

import langdetect
import openpyxl
import xlsxwriter
from nltk.corpus import wordnet

import modules.pytextrank.pytextrank.pytextrank as pyt
from modules.kku.trans.mtranslate.mtranslate import translate


class UsersSituation:
    """ Class of users situations """

    def __init__(self, data=None, category=None, result=None):
        """ Creates users situations """
        self.data = data
        self.category = category
        self.result = result
        self.translated_situation = None
        self.key_words = []

    def read_users_situation(self, filename: str) -> str:
        """ Reads users_situation from exel file. Returns users situation """
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook.active
        self.data = worksheet.cell(column=1, row=worksheet.max_row).value
        self.category = worksheet.cell(column=2, row=worksheet.max_row).value
        return self.data

    def history_user_write(self) -> None:
        """ Creates file for current user history with his situation """
        writer = xlsxwriter.Workbook("my_app/users_history.xlsx")
        worksheet = writer.add_worksheet()
        row = 0
        worksheet.write(row, 0, self.data)
        worksheet.write(row, 1, self.category)
        writer.close()

    def translate_situation(self) -> None:
        """ Translates situation into english """
        translated = translate(self.data, "en")
        self.translated_situation = translated
        return self.translated_situation

    def choicing_relevant_information_from_situation(self) -> None:
        """ Chose relevant information from situation discription """
        text = self.translated_situation
        sentence, keywords = pyt.top_keywords_sentences(text, phrase_limit=15,
                                                        sent_word_limit=150)
        for i in keywords.split():
            self.key_words.append(keywords)
        return self.key_words

    def find_synonyms(self, relevant_words_list) -> set:
        """ Returns list of synonyms and this words, that can be relevant
        to each article """
        synonyms = []
        for word in relevant_words_list:
            try:
                for syn in wordnet.synsets(word):
                    for k in syn.lemmas():
                        translated = translate(k.name(), "uk")
                        synonyms.append(translated)
            except AttributeError:
                pass
        return set(self.synonyms_check(synonyms))

    def synonyms_check(self, set_syn) -> list:
        """ Check are all word in synonyms set ukrainian """
        for word in set_syn:
            try:
                if langdetect.detect(word) == "uk":
                    pass
                else:
                    set_syn.remove(word)
            except langdetect.lang_detect_exception.LangDetectException:
                set_syn.remove(word)
        return set_syn

    def find_persons(self):
        """ Finds main persons in situation """
        pass

    def find_articles(self, user_word_set):
        """ Find articles ro situations """
        workbook = openpyxl.load_workbook("articles_with_punkts.xlsx")
        worksheet = workbook.active
        result = []
        percent = []
        counter = 1
        for i in worksheet["J"]:
            if i.value is not None and i.value != "Важливі слова":
                colu = i.value.replace(",", "")
                colu = colu.replace("'", "")
                colu = colu.replace("{", "")
                colu = colu.replace("}", "")
                colu = set(colu.split())
                lenght = len(colu.intersection(user_word_set))
                spil = float(len(colu.union(user_word_set)))*100
                percent_cur = lenght/spil
                if percent_cur != 0.0:
                    percent.append(percent_cur)
                    str_c = "A"+str(counter)
                    result.append(worksheet[str_c].value)
                counter += 1
        percent_cur = 0.0
        counter = 1
        for i in worksheet["A"]:
            if i.value is not None and i.value != "Рядки":
                index = i.value.find(".")
                colu = i.value[index+1:]
                colu = set(colu.split())
                lenght = len(colu.intersection(user_word_set))
                spil = float(len(colu.union(user_word_set))) * 100
                percent_cur = lenght / spil
                if percent_cur != 0.0:
                    percent.append(percent_cur+70)
                    str_c = "A"+str(counter)
                    result.append(worksheet[str_c].value)
                counter += 1
        if result != []:
            self.result = sorted(zip(percent, result), reverse=True)[:5]
        else:
            user_word_set = set((self.data).split())
            counter = 1
            for i in worksheet["J"]:
                if i.value is not None and i.value != "Рядки":
                    colu = i.value.replace(",", "")
                    colu = colu.replace("'", "")
                    colu = colu.replace("{", "")
                    colu = colu.replace("}", "")
                    colu = set(colu.split())
                    lenght = len(colu.intersection(user_word_set))
                    spil = float(len(colu.union(user_word_set)))*100
                    percent_cur = lenght/spil
                    if percent_cur != 0.0:
                        percent.append(percent_cur)
                        str_c = "A"+str(counter)
                        result.append(worksheet[str_c].value)
                    counter += 1
            counter = 1
            for i in worksheet["A"]:
                if i.value is not None and i.value != "Рядки":
                    index = i.value.find(".")
                    colu = i.value[index + 1:]
                    colu = set(colu.split())
                    lenght = len(colu.intersection(user_word_set))
                    spil = float(len(colu.union(user_word_set))) * 100
                    percent_cur = lenght / spil
                    if percent_cur != 0.0:
                        percent.append(percent_cur)
                        str_c = "A" + str(counter)
                        result.append(worksheet[str_c].value)
                    counter += 1
            self.result = sorted(zip(percent, result), reverse=True)[:5]
        return self.result



    def saves_result(self, filename1: str, filename2: str) -> None:
        """ Writes result to users_history and to result """
        workbook = openpyxl.load_workbook(filename1)
        worksheet = workbook.active
        worksheet.cell(column=5, row=worksheet.max_row,
                       value=str(self.result))
        workbook.save(filename1)
        workbook = openpyxl.load_workbook(filename2)
        worksheet = workbook.active
        worksheet.cell(column=2, row=worksheet.max_row,
                       value=str(self.result))
        workbook.save(filename2)
