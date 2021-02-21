""" Running web application """
import openpyxl
from flask import Flask, render_template, request

from modules.web_application.my_app.reading_users_situations.read_save_input import \
    UsersSituation

APP = Flask(__name__)


@APP.route("/")
def index():
    """ Running start.html"""
    return render_template("start.html")


@APP.route("/", methods=["GET", "POST"])
def pars_urls():
    """ Saving result from form """
    counter = 1
    situation = request.form["message"].strip()
    category = request.form["dept"]
    name = request.form["name"]
    email = request.form["email"]
    workbook = openpyxl.load_workbook(
        "result.xlsx")
    worksheet = workbook.active
    worksheet.cell(column=1, row=worksheet.max_row + 1, value=situation)
    worksheet.cell(column=2, row=worksheet.max_row, value=category)
    worksheet.cell(column=3, row=worksheet.max_row, value=name)
    worksheet.cell(column=4, row=worksheet.max_row, value=email)
    workbook.save("result.xlsx")
    user_situation = UsersSituation()
    user_situation.read_users_situation("result.xlsx")
    user_situation.translate_situation()
    users_wordset = user_situation.find_synonyms(
        user_situation.choicing_relevant_information_from_situation())
    result = user_situation.find_articles(users_wordset)
    article_1 = "Не знайдено"
    article_2 = "Не знайдено"
    article_3 = "Не знайдено"
    article_4 = "Не знайдено"
    article_5 = "Не знайдено"
    for i in range(len(result)):
        if i == 0:
            article_1 = result[i][1]
        elif i == 1:
            article_2 = result[i][1]
        elif i == 2:
            article_3 = result[i][1]
        elif i == 3:
            article_4 = result[i][1]
        elif i == 4:
            article_5 = result[i][1]
    return render_template("result.html", article_1=article_1,
                           article_2=article_2, article_3=article_3,
                           article_4=article_4, article_5=article_5)


if __name__ == "__main__":
    APP.run(debug=True)
